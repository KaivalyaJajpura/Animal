"""
Script to create a sample Excel file for bulk animal upload
This creates a template file that users can use as a reference
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Animals"

# Define header style
header_fill = PatternFill(start_color="13ec5b", end_color="13ec5b", fill_type="solid")
header_font = Font(bold=True, color="102216")
header_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add headers
headers = ["Name", "Species", "Weight (kg)", "Age (years)", "Gender"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Add sample data
sample_data = [
    ["Daisy", "Cow", 680, 5, "Female"],
    ["Bessie", "Cow", 720, 6, "Female"],
    ["Billy", "Goat", 85, 2, "Male"],
    ["Nanny", "Goat", 75, 3, "Female"],
    ["Woolly", "Sheep", 65, 3, "Male"],
    ["Fluff", "Sheep", 60, 2, "Female"],
    ["Shadow", "Buffalo", 900, 7, "Male"],
    ["Thunder", "Horse", 500, 4, "Male"],
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        if col_num == 3:  # Weight column
            cell.alignment = Alignment(horizontal="right")
        elif col_num == 4:  # Age column
            cell.alignment = Alignment(horizontal="right")

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12

# Save the workbook
wb.save('sample_animals_template.xlsx')
print("✓ Sample animals template created: sample_animals_template.xlsx")
